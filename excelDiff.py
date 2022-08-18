from tkinter import Tk
from tkinter.filedialog import askopenfilename
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, GradientFill
from openpyxl.comments import Comment

# column n번째를 알파벳으로 변경한다.
# 엑셀파일의 셀 좌표를 표시하기 위함
def columnToLetter(column_int): 
    start_index = 1 
    letter = ''
    while column_int > 25 + start_index:   
        letter += chr(65 + int((column_int-start_index)/26) - 1)
        column_int = column_int - (int((column_int-start_index)/26))*26
    letter += chr(65 - start_index + (int(column_int)))
    return letter

# 빨간 바탕색으로 칠하기
redFill = PatternFill(start_color='FFFF0000',
                      end_color='000000FF',
                      fill_type='solid')
# 흰색 -> 초록색 그라대이션 으로 바탕 칠하기
gradientFill = GradientFill(stop=("00FFFFFF", "0000FF00"))

# file path 를 dialog 창으로 선택하여 가져온다.
def getFile(msg):
    filename = ''
    while (filename == ''):
        print(msg)
        filename = askopenfilename()
        if filename == '':
            print("파일이 선택되지 않았습니다.")
    print(filename)
    return filename

def main():
    Tk().withdraw()
    originFilename = getFile("원본 파일을 선택하세요")    
    diffTargetFilename = getFile("비교할 파일을 선택하세요")    
    print("파일을 읽고 있습니다.")
    origin_file = load_workbook(originFilename)
    changed_file = load_workbook(diffTargetFilename)
    for worksheet in origin_file.sheetnames:    
        origin_sheet = origin_file[worksheet]
        # 두 file 이 공통을 갖고있는 sheet 를 대상으로 실행한다.
        if changed_file.sheetnames.__contains__(worksheet):
            print("Sheet{0} 작업 중".format(worksheet))
            changed_sheet = changed_file[worksheet]
            cnt = 0 # count for log
            for row in range(1, origin_sheet.max_row + 1):
                for col in range(1, origin_sheet.max_column + 1):
                    originCell = origin_sheet.cell(row, col)
                    changedCell = changed_sheet.cell(row, col)
                    if originCell.value != changedCell.value:
                        cnt += 1
                        memoStr = ''
                        if originCell.value == None:
                            memoStr = "Created: {1}".format(originCell.value, changedCell.value)
                        elif changedCell.value == None:
                            memoStr = "Deleted: {0}".format(originCell.value, changedCell.value)
                        else:
                            memoStr = "ChangedFrom: {0}".format(originCell.value, changedCell.value)
                        # 해당 셀이 MergedCell 일 경우 readOnly Error 가 발생한다.
                        # MergedCell 의 가장 최상단, 최좌측의 cell 만 write 가 가능하기 때문에
                        # 최상단, 최좌측의 cell 을 newTarget 으로 지정해준다
                        if type(changedCell).__name__ == 'MergedCell':
                            for mergedCell in changed_sheet.merged_cells.ranges:
                                if changedCell.coordinate in mergedCell:
                                    newTarget = changed_sheet.cell(mergedCell.min_row, mergedCell.min_col)
                                    newTarget.comment = Comment(memoStr, u'Diff Cop')
                        else :
                            changedCell.comment = Comment(memoStr, u'Diff Cop')
                        
                        changedCell.fill = gradientFill
            print("Sheet {0} 작업 완료. 변경사항 {1} 개 ".format(worksheet,cnt))
    fileExtension = diffTargetFilename.split("/")[-1].split(".")[-1]
    onlyName = diffTargetFilename.split("/")[-1].split("."+fileExtension)[0]
    exportFileName = onlyName+"_Diff_Checked.xlsx"
    # 결과 파일 export
    changed_file.save(exportFileName)
    print("{0} 파일이 생성되었습니다.".format(exportFileName))

main()